from ast import Return
from distutils.log import error
import cv2
import numpy as np
import os
from skimage.metrics import structural_similarity as ssim
from sklearn.metrics import mean_squared_error as mse
from skimage.util import random_noise
import math
from openpyxl import workbook, load_workbook, Workbook
from openpyxl.utils import get_column_letter

rank_percent = 2
approximation_numer = 20

def images_to_gray(images): # funkcja zamienia obraz rgb na obraz w odcieniach szarosci
    for i in range(len(images)):
        images[i]=cv2.cvtColor(images[i], cv2.COLOR_BGR2GRAY)
    return images

def load_images_from_folder(folder): # funkcja ładuje zdjecia z scieżki "folder"
    images = []
    for filename in os.listdir(folder):
        img = cv2.imread(os.path.join(folder,filename))
        if img is not None:
            images.append(img)
    return images

def psnr(msevalue, maxvalue): # funkcja oblicza metrykę psnr 
    value = 10*math.log(maxvalue**2/msevalue,10)
    return value

def add_noise_to_images(image_array, noise_type, specified_var="none"): # funkcja dodaje szum do zdjęć wykorzystana została funkcja random noise
    noised_images=[]                                                    # z biblioteki skimage.util
    for i in range(len(image_array)):
        if(specified_var=="none"):
            noised_images.append(random_noise(image_array[i], mode=noise_type))
        else:
            noised_images.append(random_noise(image_array[i], mode=noise_type, var=specified_var))
    return noised_images

def approximate_image(original_image, noised_image): # funkcja aproksymujaca obraz
    u, s, vt= np.linalg.svd(noised_image, full_matrices=False) # rozkład svd macierzy za pomocą funkcji svd z biblioteki numpy
    maxrank = s.size # zapamiętanie liczby wartości singularnych
    k=np.linspace(1., maxrank/rank_percent, approximation_numer) # zmienna służąca do wyznaczenia, dla których rzędów wyznaczane są aproksymacje
    k=np.rint(k) # aproksymacje wyznaczane są dla 100/rank_percent procent rzędów i wyznaczane jest approximation numer próbek
    k=k.astype(int)
    s=np.diag(s) 
    apprx=np.zeros(u.shape)
    errortable=np.zeros((4, len(k)))
    loopindex=0
    for i in k:        
        apprx=np.linalg.multi_dot((u[:,:i],s[0:i,:i],vt[:i,:])) #aproksymacja dla i rzedu oryginalnej macierzy
        apprxnew= apprx*255
        errortable[0,loopindex]=i/maxrank # zapisanie, dla jakiego procenta maxymalnego rzedu obliczane są metryki
        errortable[1,loopindex]=ssim(original_image, apprxnew, data_range=original_image.max()-original_image.min()) # metryka ssim
        errortable[2,loopindex]=mse(original_image, apprxnew) # metryka mse
        errortable[3,loopindex]=psnr(mse(original_image, apprxnew), np.amax(original_image)) # metryka psnr
        loopindex+=1
        # cv2.imshow('org', cv2.resize(original_image, (900,900))) # wyświetlanie obecnie obliczanych zdjęć, wykorzystywane przy testowaniu algorytmu
        # cv2.imshow('noise', cv2.resize(apprx, (900,900)))
        # cv2.waitKey(0)
    errortable=errortable.round(4) # zaokrąglenie wyników metryk
    return errortable # zwrócenie metryk dla aproksymacji danej macierzy

def approximation_loop(original_images, noise_type, specified_var="none"): # funkcja aproksymująca macierze dla danych szumów
    noised_images = add_noise_to_images(original_images, noise_type, specified_var) # stowrzenie macierzy obrazów zaszumionych
    original_images_iter=iter(original_images) # generator oryginalnych obrazów, oryginalne obrazy przekazywane są w celu obliczania metryk
    errortable = [] # lista, w której zapisywane są wyniki aproksymacji wszystkich macierzy dla danego szumu
    for i in noised_images:
        errortable.append(approximate_image(next(original_images_iter), i))
 
    return errortable # zwrócenie wartości metryk

def toxls(list, filename, noicename): # funkcja zapisująca dane w pliku xlsx
    wb = Workbook() # inicjalizacja pliku
    ws = wb.active
    ws.title = 'Approximation data' # tytuł pliku
    ws['A1'].value = noicename
    print(len(list))
    for i in range(len(list)): # pętla wpisująca wartości metryk 
        listiter=iter(list[i])
        ws['A'+str(2+i*5)] = 'rank %'
        ws['A'+str(3+i*5)] = 'SSIM'
        ws['A'+str(4+i*5)] = 'MSE'
        ws['A'+str(5+i*5)] = 'PSNR'
        for row in range(2+i*5,6+i*5):
            inneriter = iter(next(listiter))
            for col in range(2,2+approximation_numer):
                char = get_column_letter(col)
                ws[char + str(row)]=next(inneriter)

    wb.save(filename)

if __name__ == "__main__":
    np.set_printoptions(suppress=True)
    images = load_images_from_folder('G:/vs_code_python/SVD')
    images = images_to_gray(images)

    errortablegauss1 = approximation_loop(images, 'gaussian', 0.1**2) # wywołanie aproksymacji wszystkich zdjęć dla szumu gaussowskiego o danym odchyleniu std

    errortablegauss2 = approximation_loop(images, 'gaussian', 0.2**2) # wywołanie aproksymacji wszystkich zdjęć dla szumu gaussowskiego o danym odchyleniu std

    errortablegauss3 = approximation_loop(images, 'gaussian', 0.5**2) # wywołanie aproksymacji wszystkich zdjęć dla szumu gaussowskiego o danym odchyleniu std

    errortablespeckle1 = approximation_loop(images, 'speckle', 0.1**2) # wywołanie aproksymacji wszystkich zdjęć dla szumu speckle o danym odchyleniu std

    errortablespeckle2 = approximation_loop(images, 'speckle', 0.2**2) # wywołanie aproksymacji wszystkich zdjęć dla szumu speckle o danym odchyleniu std

    errortablespeckle3 = approximation_loop(images, 'speckle', 0.5**2) # wywołanie aproksymacji wszystkich zdjęć dla szumu speckle o danym odchyleniu std

    errortablepoisson = approximation_loop(images, 'poisson') # wywołanie aproksymacji wszystkich zdjęć dla szumu poissona

    toxls(errortablepoisson, 'poisson.xlsx', 'poisson') # zapisanie metryk odpowiednich aproksymacji do odpowiednich plików xlsx
    toxls(errortablepoisson, 'gauss1.xlsx', 'gauss1')
    toxls(errortablepoisson, 'gauss2.xlsx', 'gauss2')
    toxls(errortablepoisson, 'gauss3.xlsx', 'gauss3')
    toxls(errortablepoisson, 'speckle1.xlsx', 'speckle1')
    toxls(errortablepoisson, 'speckle2.xlsx', 'speckle2')
    toxls(errortablepoisson, 'speckle3.xlsx', 'speckle3')

    